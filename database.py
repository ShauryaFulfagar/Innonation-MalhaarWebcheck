from openpyxl import load_workbook

wb = load_workbook('Hamara Database.xlsx')


def get_details(school_no):
    try:
        ws = wb[str(school_no)]
        if ws['A2'].value == str(school_no):
            schoolDetails = [ws['A1'].value, ws['A2'].value]
            if ws['B3'].value == 'Teacher Incharge':
                teacherDetails = [ws['B1'].value, ws['B2'].value]
                student_details = []
                for col in ws.iter_cols(min_col=3):
                    col_data = []
                    for cell in col:
                        col_data.append(cell.value)
                    student_details.append(col_data)
                return schoolDetails, teacherDetails, student_details
            else:
                print('Teacher Incharge details not found for', schoolDetails[0])
        else:
            print('School details not found for', school_no)
    except:
        print('School not found for', school_no)

print(get_details("11"))

